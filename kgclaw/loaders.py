"""
Multi-format file loaders for KGClaw.

Supports: .txt, .md, .jsonl, .docx, .pdf, .html, .htm, .csv, .xlsx, .xls
Also provides recursive directory loading with auto type detection.
"""
# Copyright (c) 2026 Yanzeng Li @ BNU. MIT License.


from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any, Callable, Optional
from zipfile import ZipFile
from xml.etree import ElementTree


class LoadedDocument:
    """A document loaded from any supported format, ready for KG extraction."""

    def __init__(self, text: str, source: str, metadata: dict[str, Any] = None):
        self.text = text
        self.source = source
        self.metadata = metadata or {}

    def __repr__(self) -> str:
        return f"LoadedDocument(source={self.source!r}, chars={len(self.text)})"


# ─── Loader registry ─────────────────────────────────────────────────────────

_LOADERS: dict[str, Callable[[str], LoadedDocument]] = {}


def register_loader(extensions: list[str]):
    """Decorator to register a file loader for specific extensions."""
    def decorator(fn: Callable[[str], LoadedDocument]):
        for ext in extensions:
            _LOADERS[ext.lower()] = fn
        return fn
    return decorator


def get_loader(ext: str) -> Optional[Callable[[str], LoadedDocument]]:
    return _LOADERS.get(ext.lower())


def supported_extensions() -> list[str]:
    return sorted(_LOADERS.keys())


# ─── Plain text ──────────────────────────────────────────────────────────────

@register_loader([".txt", ".md", ".markdown", ".text"])
def load_text(path: str) -> LoadedDocument:
    p = Path(path)
    text = p.read_text(encoding="utf-8", errors="replace")
    return LoadedDocument(text=text, source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size,
    })


# ─── JSONL ───────────────────────────────────────────────────────────────────

@register_loader([".jsonl"])
def load_jsonl(path: str) -> LoadedDocument:
    p = Path(path)
    lines = []
    with open(p, encoding="utf-8", errors="replace") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                text = obj.get("data", obj.get("text", ""))
                if text:
                    lines.append(text)
            except json.JSONDecodeError:
                lines.append(line)
    return LoadedDocument(text="\n".join(lines), source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size, "lines": len(lines),
    })


# ─── DOCX ────────────────────────────────────────────────────────────────────

@register_loader([".docx"])
def load_docx(path: str) -> LoadedDocument:
    p = Path(path)
    paragraphs = []
    try:
        with ZipFile(p) as z:
            xml_content = z.read("word/document.xml")
        tree = ElementTree.fromstring(xml_content)
        ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        for p_elem in tree.iter(f"{{{ns}}}p"):
            texts = []
            for t_elem in p_elem.iter(f"{{{ns}}}t"):
                if t_elem.text:
                    texts.append(t_elem.text)
            if texts:
                paragraphs.append("".join(texts))
    except Exception:
        # fallback: read as raw text
        raw = p.read_text(encoding="utf-8", errors="replace")
        return LoadedDocument(text=raw, source=str(p), metadata={
            "filename": p.name, "ext": p.suffix, "size": p.stat().st_size, "fallback": True,
        })

    return LoadedDocument(text="\n".join(paragraphs), source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size,
        "paragraphs": len(paragraphs),
    })


# ─── PDF ─────────────────────────────────────────────────────────────────────

@register_loader([".pdf"])
def load_pdf(path: str) -> LoadedDocument:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ImportError(
            "PDF support requires pypdf. Install with: pip install pypdf>=4.0"
        )

    p = Path(path)
    reader = PdfReader(str(p))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)

    full_text = "\n\n".join(pages)
    return LoadedDocument(text=full_text, source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size,
        "pages": len(pages),
    })


# ─── HTML ────────────────────────────────────────────────────────────────────

@register_loader([".html", ".htm"])
def load_html(path: str) -> LoadedDocument:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise ImportError(
            "HTML support requires beautifulsoup4. Install with: pip install beautifulsoup4>=4.12"
        )

    p = Path(path)
    raw = p.read_text(encoding="utf-8", errors="replace")
    # Try lxml first (fastest), fall back to built-in html.parser
    try:
        soup = BeautifulSoup(raw, "lxml")
    except Exception:
        import logging
        logging.getLogger("kgclaw").warning(
            "lxml not available for HTML parsing, falling back to html.parser. "
            "Install lxml for better performance: pip install lxml"
        )
        soup = BeautifulSoup(raw, "html.parser")
    # Remove script and style
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    text = soup.get_text(separator="\n")
    # Collapse multiple blank lines
    text = re.sub(r"\n{3,}", "\n\n", text)
    return LoadedDocument(text=text.strip(), source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size,
    })


# ─── CSV ─────────────────────────────────────────────────────────────────────

@register_loader([".csv", ".tsv"])
def load_csv(path: str) -> LoadedDocument:
    p = Path(path)
    delimiter = "\t" if p.suffix == ".tsv" else ","
    rows = []
    headers = []
    try:
        with open(p, encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            for row in reader:
                rows.append(row)
    except Exception:
        # fallback: read as plain text
        text = p.read_text(encoding="utf-8", errors="replace")
        return LoadedDocument(text=text, source=str(p), metadata={
            "filename": p.name, "ext": p.suffix, "rows": 0,
        })

    # Convert to readable text representation
    lines = [f"Columns: {', '.join(headers)}", f"Total rows: {len(rows)}", ""]
    for i, row in enumerate(rows[:500]):  # cap at 500 rows
        line_parts = []
        for h in headers:
            val = row.get(h, "")
            if val:
                line_parts.append(f"{h}: {val}")
        lines.append(" | ".join(line_parts))

    text = "\n".join(lines)
    if len(rows) > 500:
        text += f"\n\n... ({len(rows) - 500} more rows)"

    return LoadedDocument(text=text, source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size,
        "rows": len(rows), "columns": headers, "is_tabular": True,
        "raw_rows": rows,  # preserve structured data for direct extraction
    })


# ─── XLSX / XLS ──────────────────────────────────────────────────────────────

@register_loader([".xlsx", ".xls"])
def load_xlsx(path: str) -> LoadedDocument:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Excel support requires openpyxl. Install with: pip install openpyxl>=3.1"
        )

    p = Path(path)
    wb = load_workbook(str(p), read_only=True, data_only=True)
    all_sheets_text = []
    all_rows_raw = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            data_rows.append({headers[i]: str(row[i]) if row[i] is not None else ""
                              for i in range(min(len(headers), len(row)))})
        all_rows_raw.extend(data_rows)

        # Text representation
        lines = [f"Sheet: {sheet_name}", f"Columns: {', '.join(headers)}",
                 f"Rows: {len(data_rows)}", ""]
        for i, dr in enumerate(data_rows[:200]):
            parts = [f"{h}: {dr.get(h, '')}" for h in headers if dr.get(h)]
            lines.append(" | ".join(parts))
        lines.append("")
        all_sheets_text.extend(lines)

    wb.close()
    text = "\n".join(all_sheets_text)
    if len(all_rows_raw) > 200:
        text += f"\n... ({len(all_rows_raw) - 200} more rows)"

    return LoadedDocument(text=text, source=str(p), metadata={
        "filename": p.name, "ext": p.suffix, "size": p.stat().st_size,
        "sheets": len(wb.sheetnames), "rows": len(all_rows_raw),
        "is_tabular": True, "raw_rows": all_rows_raw,
    })


# ─── Directory loader ────────────────────────────────────────────────────────

def load_directory(
    directory: str,
    recursive: bool = True,
    exclude_patterns: list[str] = None,
) -> list[LoadedDocument]:
    """
    Recursively load all supported files from a directory.

    Args:
        directory: Path to directory
        recursive: Whether to search subdirectories
        exclude_patterns: Glob patterns to exclude (e.g. ["*.log", ".git/**"])

    Returns:
        List of LoadedDocument objects
    """
    dir_path = Path(directory)
    if not dir_path.is_dir():
        raise NotADirectoryError(f"Not a directory: {directory}")

    exclusions = exclude_patterns or []
    docs = []

    pattern = "**/*" if recursive else "*"
    for file_path in sorted(dir_path.glob(pattern)):
        if not file_path.is_file():
            continue

        # Check exclusions
        skip = False
        for excl in exclusions:
            if file_path.match(excl):
                skip = True
                break
        if skip:
            continue

        ext = file_path.suffix.lower()
        loader = get_loader(ext)
        if loader is None:
            continue

        try:
            doc = loader(str(file_path))
            docs.append(doc)
        except Exception as e:
            # Log and skip problematic files
            import logging
            logging.getLogger("kgclaw").warning(
                f"Failed to load {file_path}: {e}"
            )

    return docs


def load_files(paths: list[str]) -> list[LoadedDocument]:
    """
    Load multiple files/directories, auto-detecting directories.

    Returns a flat list of LoadedDocument objects.
    """
    docs = []
    for p in paths:
        path = Path(p)
        if not path.exists():
            continue
        if path.is_dir():
            docs.extend(load_directory(str(path)))
        else:
            ext = path.suffix.lower()
            loader = get_loader(ext)
            if loader:
                try:
                    docs.append(loader(str(path)))
                except Exception:
                    pass
    return docs


# ─── Helper: detect structured data ──────────────────────────────────────────

def extract_structured_rows(docs: list[LoadedDocument]) -> list[dict[str, Any]]:
    """Extract raw tabular rows from all documents that have structured data."""
    rows = []
    for doc in docs:
        if doc.metadata.get("is_tabular") and "raw_rows" in doc.metadata:
            for row in doc.metadata["raw_rows"]:
                row["_source"] = doc.source
                rows.append(row)
    return rows
